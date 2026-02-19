import os
import json
import logging
from datetime import datetime
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton, WebAppInfo, ReplyKeyboardRemove
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

BOT_TOKEN = os.environ.get('BOT_TOKEN')
WEB_APP_URL = os.environ.get('WEB_APP_URL')
ADMIN_ID = 198218873

def create_xlsx_brief(data, user_info):
    """Создаёт XLSX таблицу с брифом"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Бриф"
    
    # Стили
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0F1F3D', end_color='0F1F3D', fill_type='solid')
    cell_font = Font(name='Arial', size=10)
    border = Border(
        left=Side(style='thin', color='D1D9F0'),
        right=Side(style='thin', color='D1D9F0'),
        top=Side(style='thin', color='D1D9F0'),
        bottom=Side(style='thin', color='D1D9F0')
    )
    
    # Заголовки столбцов
    headers = [
        '📅 Дата', '🆔 ID', '👤 Имя', '📧 Контакт',
        '🏢 Компания', '💼 Деятельность', '🌐 Текущий сайт',
        '🎯 Задача', '📱 Тип сайта', '📊 Страниц', '📑 Разделы',
        '👥 Аудитория', '🎯 Цели',
        '🔗 Примеры', '🎨 Стиль', '🌈 Цвета',
        '⚙️ Функции', '💻 Разработка', '📦 Материалы',
        '⏱ Сроки', '💰 Бюджет', '💭 Дополнительно'
    ]
    
    # Записываем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Данные
    now = datetime.now()
    row_data = [
        now.strftime('%d.%m.%Y %H:%M'),
        user_info.get('id', ''),
        user_info.get('name', ''),
        data.get('q20', ''),
        data.get('q1', ''),
        data.get('q2', ''),
        data.get('q4', ''),
        data.get('q3', ''),
        data.get('q5', ''),
        data.get('q6', ''),
        data.get('q7', ''),
        data.get('q8', ''),
        data.get('q9', ''),
        data.get('q10', ''),
        data.get('q11', ''),
        data.get('q12', ''),
        data.get('q13', ''),
        data.get('q14', ''),
        data.get('q15', ''),
        data.get('q16', ''),
        data.get('q17', ''),
        data.get('q18', ''),
    ]
    
    # Записываем данные
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = value if value else '—'
        cell.font = cell_font
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = border
    
    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 16  # Дата
    ws.column_dimensions['B'].width = 12  # ID
    ws.column_dimensions['C'].width = 18  # Имя
    ws.column_dimensions['D'].width = 20  # Контакт
    ws.column_dimensions['E'].width = 20  # Компания
    ws.column_dimensions['F'].width = 35  # Деятельность
    ws.column_dimensions['G'].width = 20  # Сайт
    ws.column_dimensions['H'].width = 22  # Задача
    ws.column_dimensions['I'].width = 20  # Тип сайта
    ws.column_dimensions['J'].width = 14  # Страниц
    ws.column_dimensions['K'].width = 30  # Разделы
    ws.column_dimensions['L'].width = 30  # Аудитория
    ws.column_dimensions['M'].width = 25  # Цели
    ws.column_dimensions['N'].width = 30  # Примеры
    ws.column_dimensions['O'].width = 25  # Стиль
    ws.column_dimensions['P'].width = 20  # Цвета
    ws.column_dimensions['Q'].width = 30  # Функции
    ws.column_dimensions['R'].width = 22  # Разработка
    ws.column_dimensions['S'].width = 25  # Материалы
    ws.column_dimensions['T'].width = 16  # Сроки
    ws.column_dimensions['U'].width = 20  # Бюджет
    ws.column_dimensions['V'].width = 35  # Дополнительно
    
    # Высота строк
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 80
    
    # Закрепляем первую строку
    ws.freeze_panes = 'A2'
    
    # Сохраняем в буфер
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def format_message(data, user_info):
    """Форматирует красивое сообщение с брифом"""
    msg = f"""
╔═══════════════════════════════════╗
║  📋 <b>НОВАЯ ЗАЯВКА НА ДИЗАЙН САЙТА</b>  ║
╚═══════════════════════════════════╝

👤 <b>КЛИЕНТ</b>
├ Имя: {data.get('q19', '—')}
├ Контакт: {data.get('q20', '—')}
└ ID: <code>{user_info.get('id', '—')}</code>

🏢 <b>КОМПАНИЯ</b>
├ Название: {data.get('q1', '—')}
├ Деятельность: {data.get('q2', '—')[:100]}{"..." if len(data.get('q2', '')) > 100 else ''}
└ Текущий сайт: {data.get('q4', 'Нет') if data.get('q4') else 'Нет'}

🎯 <b>ПРОЕКТ</b>
├ Задача: {data.get('q3', '—')}
├ Тип: {data.get('q5', '—')}
├ Страниц: {data.get('q6', '—')}
└ Разделы: {data.get('q7', '—')[:80]}{"..." if len(data.get('q7', '')) > 80 else ''}

👥 <b>АУДИТОРИЯ</b>
└ {data.get('q8', '—')[:120]}{"..." if len(data.get('q8', '')) > 120 else ''}

🎨 <b>ДИЗАЙН</b>
├ Стиль: {data.get('q11', '—')}
└ Цвета: {data.get('q12', 'Не указано') if data.get('q12') else 'Не указано'}

⚙️ <b>ФУНКЦИОНАЛ</b>
├ Функции: {data.get('q13', '—')[:100]}{"..." if len(data.get('q13', '')) > 100 else ''}
└ Разработка: {data.get('q14', '—')}

📦 <b>МАТЕРИАЛЫ</b>
└ {data.get('q15', '—')}

💰 <b>БЮДЖЕТ И СРОКИ</b>
├ 💵 Бюджет: <b>{data.get('q17', '—')}</b>
└ ⏱ Сроки: <b>{data.get('q16', '—')}</b>
"""
    
    # Добавляем дополнительно если есть
    if data.get('q18'):
        msg += f"\n💭 <b>ДОПОЛНИТЕЛЬНО</b>\n└ {data.get('q18')[:200]}{'...' if len(data.get('q18', '')) > 200 else ''}\n"
    
    msg += f"\n📅 Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    
    return msg

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Стартовое сообщение"""
    user = update.effective_user
    logger.info(f"🟢 {user.id} (@{user.username}) запустил бота")
    
    keyboard = [[KeyboardButton("📝 Заполнить бриф", web_app=WebAppInfo(url=WEB_APP_URL))]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    
    await update.message.reply_text(
        f"👋 Привет, {user.first_name}!\n\n"
        "Я помогу собрать информацию о вашем проекте.\n"
        "Заполните бриф — это займёт 7-10 минут.\n\n"
        "Нажмите кнопку ниже, чтобы начать 👇",
        reply_markup=reply_markup
    )

async def web_app_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка данных из квиза"""
    try:
        user = update.effective_user
        logger.info(f"🔥 ДАННЫЕ от {user.id}")
        
        raw_data = update.effective_message.web_app_data.data
        data = json.loads(raw_data)
        logger.info(f"✅ Получено {len(data)} полей")
        
        # Информация о пользователе
        user_info = {
            'id': user.id,
            'name': user.full_name,
            'username': user.username
        }
        
        # Создаём XLSX
        xlsx_buffer = create_xlsx_brief(data, user_info)
        
        # Формируем имя файла
        client_name = data.get('q19', 'Клиент').replace(' ', '-')
        now = datetime.now()
        filename = f"{now.strftime('%d.%m.%Y')}_{user.id}_{client_name}.xlsx"
        
        # Формируем сообщение
        message_text = format_message(data, user_info)
        
        # Отправляем админу
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=message_text,
            parse_mode='HTML'
        )
        
        await context.bot.send_document(
            chat_id=ADMIN_ID,
            document=xlsx_buffer,
            filename=filename,
            caption="📊 Полная информация в таблице"
        )
        
        logger.info(f"✅ Бриф отправлен админу")
        
        # Подтверждение клиенту
        await update.message.reply_text(
            "✅ <b>Отлично! Бриф получен.</b>\n\n"
            "Мы изучим вашу заявку и свяжемся с вами "
            "в ближайшее время для обсуждения деталей.\n\n"
            "Если хотите заполнить ещё один бриф — "
            "напишите /start",
            parse_mode='HTML',
            reply_markup=ReplyKeyboardRemove()
        )
        logger.info("✅ Подтверждение отправлено")
        
    except Exception as e:
        logger.error(f"❌ ОШИБКА: {e}", exc_info=True)
        await update.message.reply_text(
            "❌ Произошла ошибка при отправке.\n"
            "Попробуйте ещё раз или свяжитесь с нами напрямую.",
            reply_markup=ReplyKeyboardRemove()
        )

async def post_init(application: Application):
    """Очистка"""
    await application.bot.delete_webhook(drop_pending_updates=True)
    logger.info("✅ Готов к работе")

def main():
    """Запуск"""
    if not BOT_TOKEN or not WEB_APP_URL:
        logger.error("❌ Нет переменных окружения!")
        return
    
    logger.info(f"🚀 Запуск")
    logger.info(f"🆔 ADMIN: {ADMIN_ID}")
    logger.info(f"🌐 URL: {WEB_APP_URL}")
    
    app = Application.builder().token(BOT_TOKEN).post_init(post_init).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.StatusUpdate.WEB_APP_DATA, web_app_data))
    
    logger.info("✅ Бот готов!")
    app.run_polling(allowed_updates=Update.ALL_TYPES, drop_pending_updates=True)

if __name__ == '__main__':
    main()
